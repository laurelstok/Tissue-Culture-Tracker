#!/usr/bin/env python3
"""
TC Tracker local server
Serves index.html, handles image uploads, and generates Excel exports.
Run with: python3 server.py
Then open: http://localhost:8080
"""

import os, cgi, json, shutil, io
from http.server import HTTPServer, SimpleHTTPRequestHandler
from urllib.parse import urlparse

UPLOAD_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'images')
os.makedirs(UPLOAD_DIR, exist_ok=True)


def parse_sci(s):
    if not s: return None
    s = str(s).strip().replace(',','')
    try: return float(s)
    except: pass
    import re
    m = re.match(r'^([\d.]+)?\xd710([^\d]*)$', s)
    if not m: m = re.match(r'^([\d.]+)?\u00d710([^\d]*)$', s)
    if m:
        sup = {'\u2070':0,'\u00b9':1,'\u00b2':2,'\u00b3':3,'\u2074':4,
               '\u2075':5,'\u2076':6,'\u2077':7,'\u2078':8,'\u2079':9,'\u207b':'-'}
        coeff = float(m.group(1)) if m.group(1) else 1.0
        exp_str = ''.join(str(sup.get(c,c)) for c in m.group(2))
        try: return coeff * (10 ** int(exp_str))
        except: return None
    return None


def get_lineage_project(rec, passages):
    seen, cur = set(), rec
    while cur and cur.get('id') not in seen:
        seen.add(cur.get('id'))
        if cur.get('project'): return cur['project']
        pid = cur.get('parent')
        if not pid: break
        cur = next((p for p in passages if p.get('id') == pid), None)
    return ''


def calc_fold_change(rec, passages):
    total = parse_sci(rec.get('totalViableCells'))
    if not total: return None
    seeded = parse_sci(rec.get('seedingTotal'))
    if not seeded and rec.get('vessels'):
        s = sum((parse_sci(v.get('seedingTotal')) or 0)*(v.get('qty') or 1) for v in rec['vessels'])
        if s: seeded = s
    if not seeded:
        par = next((p for p in passages if p.get('id') == rec.get('parent')), None)
        if par:
            seeded = parse_sci(par.get('seedingTotal'))
            if not seeded and par.get('vessels'):
                s = sum((parse_sci(v.get('seedingTotal')) or 0)*(v.get('qty') or 1) for v in par['vessels'])
                if s: seeded = s
            if not seeded and par.get('plateData'):
                wa = {'6-well plate':9.5,'12-well plate':3.8,'24-well plate':1.9,
                      '48-well plate':0.95,'96-well plate':0.32,'384-well plate':0.056}
                area = wa.get(par.get('wells',''), 9.5)
                tw = 0
                for plate in par['plateData'].values():
                    for well in plate.values():
                        if well.get('count'): tw += parse_sci(well['count']) or 0
                        elif well.get('seeding') and well.get('occupied'): tw += (parse_sci(well['seeding']) or 0) * area
                if tw: seeded = tw
    return (total / seeded) if seeded and seeded > 0 else None


def build_excel(data):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    passages = data.get('passages', [])
    projects = data.get('projects', [])

    HEADER_BG = 'FF1A1915'
    ABG = {'thaw':'FFE6F1FB','inherit':'FFEEEDFE','passage':'FFE1F5EE',
           'freeze':'FFFAEEDA','experiment':'FFFAECE7'}
    thin = Side(style='thin', color='FFD0D0D0')
    bdr  = Border(left=thin, right=thin, top=thin, bottom=thin)

    def hcell(ws, r, c, v, w=None):
        cell = ws.cell(r, c, v)
        cell.font = Font(bold=True, color='FFFFFFFF', name='Arial', size=10)
        cell.fill = PatternFill('solid', start_color=HEADER_BG)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = bdr
        if w: ws.column_dimensions[get_column_letter(c)].width = w
        return cell

    def dcell(ws, r, c, v, bg=None, align='left', fmt=None):
        cell = ws.cell(r, c, v)
        cell.font = Font(name='Arial', size=10)
        if bg: cell.fill = PatternFill('solid', start_color=bg)
        cell.alignment = Alignment(horizontal=align, vertical='center', wrap_text=True)
        cell.border = bdr
        if fmt: cell.number_format = fmt
        return cell

    def vstr(rec):
        if rec.get('vessels'):
            return ', '.join('({}){}'.format(v.get('qty',1), v.get('type','')) for v in rec['vessels'])
        return rec.get('wells', '')

    # Sheet 1: Records
    ws1 = wb.active
    ws1.title = 'Records'
    ws1.freeze_panes = 'A2'
    ws1.row_dimensions[1].height = 36
    c1 = [
        ('Record ID',14),('Parent ID',14),('Line',14),('Project',16),
        ('Action',10),('P#',6),('Date',12),('Days',6),
        ('Condition',20),('Substrate',14),('Vessel(s)',22),
        ('Split ratio',10),('Confluency %',10),
        ('Seeding density',14),('Total seeded',14),
        ('Viable cells/mL',14),('% Viability',10),('Volume mL',10),
        ('Total viable cells',16),('Fold change',12),
        ('Vials',8),('Cells/vial',14),('Cryoprotectant',16),
        ('Storage',18),('Vial viability %',12),
        ('Exp ID',12),('Assay',16),('Treatment',20),('Timepoints',14),
        ('Operator',14),('Notes',30),
        ('Feeds',8),('Images',8),
    ]
    for ci,(label,w) in enumerate(c1,1): hcell(ws1,1,ci,label,w)

    for ri,rec in enumerate(passages,2):
        bg = ABG.get(rec.get('action',''))
        proj = get_lineage_project(rec, passages)
        fc = calc_fold_change(rec, passages)
        vals = [
            rec.get('id',''), rec.get('parent','') or '', rec.get('line',''), proj,
            rec.get('action',''), rec.get('passageNum',''), rec.get('date',''),
            rec.get('days','') or '', rec.get('condition',''), rec.get('substrate',''),
            vstr(rec), rec.get('splitRatio',''), rec.get('confluency','') or '',
            rec.get('seedingDensity',''), rec.get('seedingTotal',''),
            rec.get('viableCellsPerMl',''), rec.get('viabilityPct','') or '',
            rec.get('volumeMl','') or '', rec.get('totalViableCells',''),
            round(fc,2) if fc else '',
            rec.get('vials','') or '', rec.get('cellsPerVial',''),
            rec.get('cryo',''), rec.get('storage',''), rec.get('viability','') or '',
            rec.get('expId',''), rec.get('assay',''), rec.get('treatment',''),
            rec.get('timepoints',''), rec.get('user',''), rec.get('note',''),
            len(rec.get('feeds',[])), len(rec.get('images',[])),
        ]
        for ci,v in enumerate(vals,1):
            align = 'center' if ci in (6,8,12,13,20,32,33) else 'left'
            fmt = '0.00"x"' if ci == 20 and v else None
            dcell(ws1, ri, ci, v, bg=bg, align=align, fmt=fmt)

    ws1.auto_filter.ref = ws1.dimensions

    # Sheet 2: Feed log
    ws2 = wb.create_sheet('Feed log')
    ws2.freeze_panes = 'A2'
    c2 = [('Record ID',14),('Line',14),('Project',14),('Action',10),
          ('Feed date',12),('Media type',22),('Volume mL',10),
          ('Operator',14),('Notes',30),('Cat #',14),('Lot #',14),('Expiration',12)]
    for ci,(l,w) in enumerate(c2,1): hcell(ws2,1,ci,l,w)
    row = 2
    for rec in passages:
        proj = get_lineage_project(rec, passages)
        bg = ABG.get(rec.get('action',''))
        for feed in rec.get('feeds',[]):
            vals = [rec.get('id',''),rec.get('line',''),proj,rec.get('action',''),
                    feed.get('date',''),feed.get('media',''),feed.get('volumeMl','') or '',
                    feed.get('user',''),feed.get('note',''),
                    feed.get('mediaCat',''),feed.get('mediaLot',''),feed.get('mediaExp','')]
            for ci,v in enumerate(vals,1): dcell(ws2,row,ci,v,bg=bg)
            row += 1
    if row == 2: ws2.cell(2,1,'No feed events recorded yet.')

    # Sheet 3: Plate wells
    ws3 = wb.create_sheet('Plate wells')
    ws3.freeze_panes = 'A2'
    c3 = [('Record ID',14),('Line',14),('Project',14),('Action',10),('Date',12),
          ('Plate key',14),('Well',8),('Seeding density',14),('Cell count',14),
          ('Condition',20),('Treatment',20),('Notes',30),('Contaminated',12)]
    for ci,(l,w) in enumerate(c3,1): hcell(ws3,1,ci,l,w)
    row = 2
    for rec in passages:
        proj = get_lineage_project(rec, passages)
        bg = ABG.get(rec.get('action',''))
        for pk,plate in rec.get('plateData',{}).items():
            for wid,well in plate.items():
                if not well.get('occupied') and not well.get('contaminated'): continue
                vals = [rec.get('id',''),rec.get('line',''),proj,rec.get('action',''),
                        rec.get('date',''),pk,wid,
                        well.get('seeding',''),well.get('count',''),
                        well.get('cond',''),well.get('treat',''),well.get('note',''),
                        'Yes' if well.get('contaminated') else '']
                for ci,v in enumerate(vals,1): dcell(ws3,row,ci,v,bg=bg)
                row += 1
    if row == 2: ws3.cell(2,1,'No plate well data recorded yet.')

    # Sheet 4: Summary by cell line
    ws4 = wb.create_sheet('Summary')
    ws4.freeze_panes = 'A2'
    c4 = [('Cell line',16),('Project(s)',22),('Records',10),
          ('Max passage',12),('Avg days/passage',16),
          ('Records with fold change',20),('Avg fold change',16),
          ('Freeze stocks',12),('Experiments',12)]
    for ci,(l,w) in enumerate(c4,1): hcell(ws4,1,ci,l,w)
    lines = list(dict.fromkeys(r.get('line','') for r in passages))
    for ri,line in enumerate(lines,2):
        lp = [r for r in passages if r.get('line')==line]
        projs = list(dict.fromkeys(filter(None,(get_lineage_project(r,passages) for r in lp))))
        days = [r.get('days',0) for r in lp if r.get('days',0)>0]
        avg_d = round(sum(days)/len(days),1) if days else ''
        max_p = max((r.get('passageNum',0) for r in lp),default=0)
        fcs = [f for f in (calc_fold_change(r,passages) for r in lp) if f is not None]
        avg_fc = round(sum(fcs)/len(fcs),2) if fcs else ''
        vals = [line,', '.join(projs),len(lp),max_p,avg_d,len(fcs),avg_fc,
                sum(1 for r in lp if r.get('action')=='freeze'),
                sum(1 for r in lp if r.get('action')=='experiment')]
        for ci,v in enumerate(vals,1):
            dcell(ws4,ri,ci,v,align='center' if ci>2 else 'left',
                  fmt='0.00"x"' if ci==7 and v else None)

    # Sheet 5: Projects
    if projects:
        ws5 = wb.create_sheet('Projects')
        hcell(ws5,1,1,'Project name',20)
        for ri,p in enumerate(projects,2): dcell(ws5,ri,1,p)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


class TCHandler(SimpleHTTPRequestHandler):

    def do_POST(self):
        parsed = urlparse(self.path)
        if parsed.path == '/upload':
            self.handle_upload()
        elif parsed.path == '/export-excel':
            self.handle_excel_export()
        else:
            self.send_error(404, 'Not found')

    def handle_excel_export(self):
        try:
            length = int(self.headers.get('Content-Length', 0))
            body = self.rfile.read(length)
            data = json.loads(body)
            xlsx = build_excel(data)
            self.send_response(200)
            self.send_header('Content-Type',
                'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            self.send_header('Content-Disposition', 'attachment; filename="tc-records.xlsx"')
            self.send_header('Content-Length', str(len(xlsx)))
            self.send_header('Access-Control-Allow-Origin', '*')
            self.end_headers()
            self.wfile.write(xlsx)
            print('  [excel] exported {} records'.format(len(data.get('passages',[]))))
        except Exception as e:
            import traceback; traceback.print_exc()
            self.send_json(500, {'error': str(e)})

    def handle_upload(self):
        try:
            ct = self.headers.get('Content-Type', '')
            if 'multipart/form-data' not in ct:
                self.send_json(400, {'error': 'Expected multipart/form-data'}); return
            form = cgi.FieldStorage(fp=self.rfile, headers=self.headers,
                environ={'REQUEST_METHOD':'POST','CONTENT_TYPE':ct})
            if 'file' not in form:
                self.send_json(400, {'error': 'No file field'}); return
            fi = form['file']
            fn = os.path.basename(fi.filename)
            fn = ''.join(c for c in fn if c.isalnum() or c in '._- ').strip()
            if not fn: self.send_json(400, {'error': 'Invalid filename'}); return
            if 'filename' in form:
                cn = ''.join(c for c in os.path.basename(form['filename'].value) if c.isalnum() or c in '._- ')
                if cn: fn = cn
            dest = os.path.join(UPLOAD_DIR, fn)
            with open(dest,'wb') as f: shutil.copyfileobj(fi.file, f)
            size = os.path.getsize(dest)
            print('  [upload] {} ({:,} bytes)'.format(fn, size))
            self.send_json(200, {'filename': fn, 'size': size})
        except Exception as e:
            print('  [upload error]', e)
            self.send_json(500, {'error': str(e)})

    def send_json(self, code, data):
        body = json.dumps(data).encode()
        self.send_response(code)
        self.send_header('Content-Type', 'application/json')
        self.send_header('Content-Length', str(len(body)))
        self.send_header('Access-Control-Allow-Origin', '*')
        self.end_headers()
        self.wfile.write(body)

    def do_OPTIONS(self):
        self.send_response(200)
        self.send_header('Access-Control-Allow-Origin', '*')
        self.send_header('Access-Control-Allow-Methods', 'GET, POST, OPTIONS')
        self.send_header('Access-Control-Allow-Headers', 'Content-Type')
        self.end_headers()

    def log_message(self, fmt, *args):
        if args and str(args[1]) not in ('200','304'):
            super().log_message(fmt, *args)


if __name__ == '__main__':
    port = 8080
    server = HTTPServer(('localhost', port), TCHandler)
    print('TC Tracker running at http://localhost:{}'.format(port))
    print('Images folder: {}'.format(UPLOAD_DIR))
    print('Press Ctrl+C to stop\n')
    try:
        server.serve_forever()
    except KeyboardInterrupt:
        print('\nStopped.')